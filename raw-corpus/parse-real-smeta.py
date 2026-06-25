#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
parse-real-smeta.py — извлекает РЕАЛЬНЫЕ расценки из реальной локальной сметы РК
(экспорт сметной программы, Форма 4 ресурсная по НДЦС РК 8.01-08-2022).

Источник: .xlsx с листами вида <код>_Q9 (Форма 4 — локальная смета).
Раскладка строк (проверено на примере «коллектор»):
  заголовок данных: r с «номер по порядку» + «обоснование»
  колонки данных:  C0 № | C1 обоснование/код | C2 наименование | C3 ед.изм |
                   C4 кол-во | C5 цена ед., ₸ | C6 общая стоимость, ₸
  главная позиция: C0 — целое; C1 — код норматива (ЭСН РК / ССЦ / перевозка)
  ресурсы позиции: C0 вида «1.1.1»; C1 — код ресурса; C3 — ед. (чел.-ч/маш.-ч/…)

ЗАЩИТА ОТ ИНЪЕКЦИЙ ВЫВОДА: в выгрузке встречаются подменённые/битые строки.
Принимаем позицию ТОЛЬКО если арифметика сходится: |C4*C5 - C6| <= EPS.
Это надёжно отсекает фабрикованные строки (литеральные «...» и пр.).

Выход: frontend/src/app/smeta-trainer/data/real-rates.json
Запуск: python -X utf8 raw-corpus/parse-real-smeta.py "<path.xlsx>" [out.json]
"""
import sys, json, re, os

try:
    import openpyxl
except ImportError:
    sys.exit("Нужен openpyxl: python -m pip install openpyxl")

EPS = 2.0  # ₸, допуск округления

CODE_RE = re.compile(r"\d{3,4}-\d{2,4}-\d{2,4}")          # 1101-0705-0107 / 3414-104-0501
RES_CODE_RE = re.compile(r"\d{2,4}-\d{2,4}(?:-\d{2,4})?")  # 001-0138 (труд) / 343-102-0401
ESN_DOC_RE = re.compile(r"ЭСН\s*РК\s*[\d.\-]+", re.I)
SUBNUM_RE = re.compile(r"^\d+\.\d+(\.\d+)+$")             # 1.1.1 — строка ресурса


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        t = v.replace("\xa0", "").replace(" ", "").replace(",", ".")
        try:
            return float(t)
        except ValueError:
            return None
    return None


def text(v):
    return str(v).strip() if v is not None else ""


def classify_basis(s):
    low = s.lower()
    if "эсн" in low:
        return "ЭСН РК"
    if "ссц" in low:
        return "ССЦ РК"
    if "перевоз" in low:
        return "Перевозка"
    return "прочее"


def classify_kind(unit, name):
    u = unit.lower()
    nl = name.lower()
    if "чел" in u or "труд" in nl:
        return "труд"
    if "маш" in u or "механизм" in nl:
        return "машины"
    if u in ("т", "т*км", "ткм") or "перевоз" in nl:
        return "перевозка"
    return "материал"


def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    obj = smeta_no = priceLevel = None
    header_ri = None
    totals = {}
    for ri, row in enumerate(rows):
        if not row:
            continue
        c0 = text(row[0])
        joined = " ".join(text(c) for c in row if c is not None)
        low = joined.lower()
        if "наименование объекта" in low and len(row) > 2:
            obj = text(row[2])
        m = re.search(r"локальная смета\s*[№N\?]?\s*([\d\-]+)", low)
        if m:
            smeta_no = m.group(1)
        if "составлен" in low and "цен" in low:
            priceLevel = joined.strip()[:80]
        if "номер по порядку" in low and "обоснование" in low:
            header_ri = ri
        lbl = text(row[2]).lower() if len(row) > 2 else ""
        tcol = num(row[6]) if len(row) > 6 else None
        if lbl.startswith("всего по смете"):
            totals["всего"] = tcol
        elif lbl.startswith("затраты на труд"):
            totals["труд"] = tcol
        elif lbl.startswith("машины и механизм"):
            totals["машины"] = tcol
        elif lbl.startswith("перевозки"):
            totals["перевозки"] = tcol

    positions = []
    rejected = 0
    start = (header_ri + 1) if header_ri is not None else 0
    cur = None  # текущая главная позиция (для привязки ресурсов)

    for row in rows[start:]:
        if not row or len(row) < 7:
            continue
        c0 = text(row[0])
        basis = text(row[1])

        # --- ресурс позиции (1.1.1) ---
        if cur is not None and SUBNUM_RE.match(c0) and RES_CODE_RE.search(basis):
            unit = text(row[3])
            qty = num(row[4])
            price = num(row[5])
            tot = num(row[6])
            if None not in (qty, price, tot) and abs(qty * price - tot) <= EPS and qty > 0:
                cur["resources"].append({
                    "code": RES_CODE_RE.search(basis).group(0),
                    "kind": classify_kind(unit, text(row[2])),
                    "name": text(row[2]),
                    "qtyPerUnit": qty,
                    "unit": unit,
                    "basePrice": price,
                    "total": tot,
                })
            continue

        # --- главная позиция (целое в C0) ---
        n = num(c0)
        if n is None or n != int(n):
            continue
        if not CODE_RE.search(basis):
            continue
        qty = num(row[4])
        price = num(row[5])
        tot = num(row[6])
        if None in (qty, price, tot):
            rejected += 1
            continue
        if abs(qty * price - tot) > EPS:        # арифметический гейт
            rejected += 1
            continue
        doc = ESN_DOC_RE.search(basis)
        cur = {
            "n": int(n),
            "code": CODE_RE.search(basis).group(0),
            "basisDoc": doc.group(0) if doc else None,
            "basisClass": classify_basis(basis),
            "name": text(row[2]),
            "unit": text(row[3]),
            "qty": qty,
            "unitPrice": price,
            "total": tot,
            "resources": [],
        }
        positions.append(cur)

    return obj, smeta_no, priceLevel, totals, positions, rejected


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else None
    if not src or not os.path.exists(src):
        sys.exit("Укажи путь к .xlsx первым аргументом")
    here = os.path.dirname(os.path.abspath(__file__))
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        here, "..", "frontend", "src", "app", "smeta-trainer", "data", "real-rates.json")

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    q9 = [n for n in wb.sheetnames if n.endswith("_Q9")]
    smety = []
    total_pos = total_res = total_rej = 0
    for name in q9:
        obj, smeta_no, priceLevel, totals, positions, rejected = parse_sheet(wb[name])
        total_pos += len(positions)
        total_res += sum(len(p["resources"]) for p in positions)
        total_rej += rejected
        if positions:
            smety.append({
                "sheet": name,
                "smetaNo": smeta_no,
                "object": obj,
                "priceLevel": priceLevel,
                "totals": totals,
                "positions": positions,
            })

    result = {
        "version": "real-2026.1",
        "source": os.path.basename(src),
        "norm": "НДЦС РК 8.01-08-2022, Форма 4 (ресурсная локальная смета)",
        "smetyCount": len(smety),
        "positionsCount": total_pos,
        "resourcesCount": total_res,
        "smety": smety,
    }
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=1)

    print("Q9 sheets:", len(q9))
    print("smety with data:", len(smety))
    print("positions accepted:", total_pos, "| resources:", total_res,
          "| rejected (arith gate):", total_rej)
    print("out:", os.path.relpath(out, here))


if __name__ == "__main__":
    main()
